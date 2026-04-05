from .base import BaseParser, ParsedDocument
import io
import csv


class XLSXParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> ParsedDocument:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            text = ""

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"\n=== Sheet: {sheet_name} ===\n"

                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    text += " | ".join(row_data) + "\n"

            return ParsedDocument(
                texto=text,
                estrutura=self._detect_structure(text),
                metadados={
                    "arquivo": filename,
                    "tamanho": len(content),
                    "folhas": len(wb.sheetnames),
                },
                alertas=[],
                limitacoes=[],
            )
        except Exception as e:
            result = self._create_empty_result(filename)
            result.alertas.append(f"Erro ao processar XLSX: {str(e)}")
            return result


class CSVParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> ParsedDocument:
        try:
            text = content.decode("utf-8")
            lines = text.split("\n")
            reader = csv.reader(lines)

            formatted = ""
            for row in reader:
                formatted += " | ".join(row) + "\n"

            return ParsedDocument(
                texto=formatted,
                estrutura=self._detect_structure(formatted),
                metadados={
                    "arquivo": filename,
                    "tamanho": len(content),
                    "linhas": len(lines),
                },
                alertas=[],
                limitacoes=[],
            )
        except Exception as e:
            try:
                text = content.decode("latin-1")
                lines = text.split("\n")
                formatted = ""
                for row in lines:
                    formatted += row + "\n"

                return ParsedDocument(
                    texto=formatted,
                    estrutura=self._detect_structure(formatted),
                    metadados={
                        "arquivo": filename,
                        "tamanho": len(content),
                        "encoding": "latin-1",
                        "linhas": len(lines),
                    },
                    alertas=["Arquivo codificado em latin-1"],
                    limitacoes=[],
                )
            except Exception as fallback_error:
                result = self._create_empty_result(filename)
                result.alertas.append(
                    f"Erro ao processar CSV: {str(e)} - Fallback também falhou: {str(fallback_error)}"
                )
                return result
